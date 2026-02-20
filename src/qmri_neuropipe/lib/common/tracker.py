import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
import logging
import os
import time
import shutil

class NeuroimagingTracker:
    """
    Manages a multi-sheet Excel file for tracking neuroimaging data,
    processing status, and quality metrics.
    """
    
    CORE_SHEETS = [
        'Summary', 'Subject_Metadata', 'Processing_Status', 
        'Anatomical_Status', 'Diffusion_Status', 'Relaxometry_Status',
        'Quality_Metrics', 'Data_Files', 
        'Software_Versions', 'Alert_History'
    ]

    @staticmethod
    def create_empty_tracker(path: Path):
        """Create a new Excel file with the standard tracking structure."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        sheets = {
            'Summary': ['Metric', 'Value'],
            'Subject_Metadata': ['Subject_ID', 'Session', 'Study', 'Age', 'Sex', 'Group', 'Scan_Date'],
            'Processing_Status': ['Subject_ID', 'Session', 'Study', 'Overall_Pipeline_Status', 'Last_Processing_Date'],
            'Anatomical_Status': ['Subject_ID', 'Session', 'Study', 'Denoising', 'Gibbs_Ringing', 'Reorienting', 'Bias_Correction', 'Brain_Masking', 'Segmentation', 'Segmentation_Method', 'Coregistration', 'Analysis', 'Atlases', 'Overall_Status', 'Last_Update'],
            'Diffusion_Status': ['Subject_ID', 'Session', 'Study', 'Denoising', 'Gibbs_Ringing', 'Eddy_Correction', 'Bias_Correction', 'Topup', 'SynB0', 'Coregistration', 'Reorienting', 'Model_Fits', 'Atlases', 'Overall_Status', 'Last_Update'],
            'Relaxometry_Status': ['Subject_ID', 'Session', 'Study', 'Denoising', 'Gibbs_Correction', 'Motion_Correction', 'B1_Mapping_Method', 'Analysis', 'Overall_Status', 'Last_Update'],
            'Volume_Statistics': ['Subject_ID', 'Session', 'Study', 'Method', 'Structure', 'Volume_mm3', 'ICV_Normalized'],
            'ROI_Metrics': ['Subject_ID', 'Session', 'Study', 'ROI_Source', 'ROI_Name', 'Modality', 'Metric', 'Statistic', 'Value'],
            'Quality_Metrics': ['Subject_ID', 'Session', 'Study', 'Motion_FD_Mean', 'DWI_SNR'],
            'Data_Files': ['Subject_ID', 'Session', 'Study', 'T1w_Present', 'DWI_Present'],
            'Software_Versions': ['Subject_ID', 'Session', 'Study', 'Pipeline_Version'],
            'Alert_History': ['Alert_ID', 'Subject_ID', 'Study', 'Alert_Type', 'Alert_Status']
        }
        
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for name, cols in sheets.items():
                pd.DataFrame(columns=cols).to_excel(writer, sheet_name=name, index=False)
                
            # Add README
            sheet_descriptions = {
                'Summary': 'High-level study summary',
                'Subject_Metadata': 'Subject demographics and metadata',
                'Processing_Status': 'Overall pipeline status',
                'Anatomical_Status': 'Anatomical processing details',
                'Diffusion_Status': 'Diffusion processing details',
                'Relaxometry_Status': 'Relaxometry processing details',
                'Volume_Statistics': 'Anatomical structure volumes (FSL FAST/FIRST/FreeSurfer)',
                'ROI_Metrics': 'Cross-modal ROI metrics (diffusion/relaxometry in anatomical ROIs)',
                'Quality_Metrics': 'Quality control metrics',
                'Data_Files': 'Input/Output file paths',
                'Software_Versions': 'Software and pipeline versions',
                'Alert_History': 'History of study alerts',
                'README': 'Sheet documentation'
            }
            readme_data = []
            for s in sheets.keys():
                readme_data.append({'Sheet': s, 'Description': sheet_descriptions.get(s, 'Data sheet')})
            
            pd.DataFrame(readme_data).to_excel(writer, sheet_name='README', index=False)
        
        return NeuroimagingTracker(path)

    def __init__(self, excel_path: Path, logger: Optional[logging.Logger] = None, auto_save: bool = True):
        self.excel_path = Path(excel_path)
        self.logger = logger or logging.getLogger("Tracker")
        self.auto_save = auto_save
        self._data: Dict[str, pd.DataFrame] = {}
        
        if self.excel_path.exists():
            self.load()

    def load(self):
        """Load all sheets from the Excel file."""
        if not self.excel_path.exists():
            return
            
        if self.excel_path.stat().st_size == 0:
            self.logger.warning(f"Tracker file {self.excel_path} is empty. Skipping load.")
            return

        # Acquire lock for reading too, just in case someone is writing
        lock_path = self.excel_path.with_suffix(self.excel_path.suffix + ".lock")
        try:
            self._acquire_lock(lock_path)
            # Use openpyxl engine explicitly for better multi-sheet support
            with pd.ExcelFile(self.excel_path, engine='openpyxl') as xls:
                for sheet_name in xls.sheet_names:
                    self._data[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
            self.logger.info(f"Loaded tracker from {self.excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to load tracker: {e}")
            raise
        finally:
            self._release_lock(lock_path)

    def _acquire_lock(self, lock_path: Path, timeout: int = 60):
        """
        Robust file-based lock for multi-process safety.
        Writes current PID to the lock file and detects stale locks.
        """
        start_time = time.time()
        pid = os.getpid()
        
        while True:
            try:
                # Try to create the lock file
                fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                with os.fdopen(fd, 'w') as f:
                    f.write(str(pid))
                return True
            except FileExistsError:
                # Check for stale lock
                try:
                    with open(lock_path, 'r') as f:
                        lock_pid = int(f.read().strip())
                    
                    # Check if process is still running
                    if not self._pid_exists(lock_pid):
                        self.logger.warning(f"Detected stale lock from PID {lock_pid}. Removing it.")
                        self._release_lock(lock_path)
                        continue # Try again immediately
                except (ValueError, OSError, PermissionError):
                    # If we can't read it, assume it's being written or locked by someone else
                    pass

                if time.time() - start_time > timeout:
                    self.logger.warning(f"Timeout waiting for tracker lock on {lock_path}")
                    return False
                time.sleep(0.5)

    def _pid_exists(self, pid: int) -> bool:
        """Check if a process ID exists."""
        if pid <= 0: return False
        try:
            os.kill(pid, 0)
        except ProcessLookupError:
            return False
        except PermissionError:
            return True # Exists but no permission
        else:
            return True

    def _release_lock(self, lock_path: Path):
        """Release the file-based lock."""
        if lock_path.exists():
            try:
                # Optional: check if we own the lock before removing?
                # For now, simple remove is fine as we only release in finally blocks or on stale detection
                os.remove(lock_path)
            except OSError:
                pass

    def save(self, force: bool = False):
        """
        Save all data back to the multi-sheet Excel file.
        Uses atomic replace and lightweight backups to prevent corruption.
        """
        if not self._data:
            return
            
        if not self.auto_save and not force:
            return

        lock_path = self.excel_path.with_suffix(self.excel_path.suffix + ".lock")
        temp_path = self.excel_path.with_suffix(self.excel_path.suffix + ".tmp")
        bak_path = self.excel_path.with_suffix(self.excel_path.suffix + ".bak")

        try:
            # Create directory if it doesn't exist
            self.excel_path.parent.mkdir(parents=True, exist_ok=True)
            
            if self._acquire_lock(lock_path):
                # Before saving, re-load to catch changes from other processes!
                if self.excel_path.exists() and self.excel_path.stat().st_size > 0:
                    try:
                        # CREATE BACKUP BEFORE MODIFYING
                        # Use copyfile instead of copy to avoid metadata permission issues on shared drives
                        if bak_path.exists():
                            try: bak_path.unlink()
                            except: pass
                        shutil.copyfile(self.excel_path, bak_path)

                        with pd.ExcelFile(self.excel_path, engine='openpyxl') as xls:
                            for sheet_name in xls.sheet_names:
                                try:
                                    existing_df = pd.read_excel(xls, sheet_name=sheet_name)
                                    if sheet_name in ['Summary', 'README']:
                                        continue

                                    if sheet_name in self._data:
                                        # UPSERT: Merge existing with current
                                        pk_cols = [
                                            'Subject_ID', 'Session', 'Study', 'Modality', 'Step_Name', 'Parameter',
                                            'Atlas', 'ROI_Source', 'ROI_Name', 'Metric', 'Statistic', 'Model', 
                                            'Structure', 'Method', 'Alert_ID'
                                        ]
                                        subset = [c for c in pk_cols if c in existing_df.columns and c in self._data[sheet_name].columns]
                                        
                                        # Filter out empty DataFrames to avoid FutureWarning
                                        dfs_to_concat = [d for d in [existing_df, self._data[sheet_name]] if not d.empty]
                                        if dfs_to_concat:
                                            merged = pd.concat(dfs_to_concat, ignore_index=True)
                                            if subset:
                                                # Normalize PK columns to string for consistent comparison
                                                for pk_col in subset:
                                                    if pk_col in merged.columns:
                                                        merged[pk_col] = merged[pk_col].astype(str).str.strip()
                                                        # Normalize session values: strip ses- prefix and leading zeros
                                                        if pk_col == 'Session':
                                                            def _norm_session(x):
                                                                if pd.isna(x) or x is None or x == 'nan': return 'N/A'
                                                                s = str(x).strip().replace('ses-', '')
                                                                # Strip leading zeros for numeric comparison
                                                                if s.isdigit(): s = str(int(s))
                                                                return s if s else 'N/A'
                                                            merged[pk_col] = merged[pk_col].apply(_norm_session)
                                                self._data[sheet_name] = merged.drop_duplicates(subset=subset, keep='last')
                                            else:
                                                # If no PKs found but we have memory data, we must decide:
                                                # For generated/calculated sheets, memory usually wins.
                                                # For safety, if it's a known tracking sheet, we keep the newest.
                                                self._data[sheet_name] = merged.drop_duplicates(keep='last')
                                        else:
                                            pass
                                    elif sheet_name == 'README':
                                        pass
                                    else:
                                        # Even if not in current _data, preserve existing
                                        self._data[sheet_name] = existing_df
                                except Exception as e_sheet:
                                    self.logger.warning(f"Error merging sheet {sheet_name}: {e_sheet}")
                    except Exception as e:
                        self.logger.error(f"FATAL: Tracker corruption detected in {self.excel_path}. Aborting save to prevent data loss. Original error: {e}")
                        self.logger.info(f"Please restore from backup: {bak_path}")
                        return

                # ATOMIC SAVE via temp file
                try:
                    with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                        for sheet_name, df in self._data.items():
                            # Sorting: Subject_ID, Session
                            if isinstance(df, pd.DataFrame) and 'Subject_ID' in df.columns:
                                sort_cols = ['Subject_ID']
                                if 'Session' in df.columns: sort_cols.append('Session')
                                if 'Study' in df.columns: sort_cols.append('Study')
                                # For Metrics sheets, also sort by Atlas/ROI
                                if 'Metrics' in sheet_name:
                                     if 'Atlas' in df.columns: sort_cols.append('Atlas')
                                     if 'ROI_Name' in df.columns: sort_cols.append('ROI_Name')
                                
                                df = df.sort_values(by=sort_cols)
                                
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                        # Recalculate summary and apply styles
                        self._recalculate_summary()
                        # Re-write summary sheet since it's updated
                        if 'Summary' in self._data:
                             self._data['Summary'].to_excel(writer, sheet_name='Summary', index=False)
                             
                        workbook = writer.book
                        self._apply_styles(workbook)
                        self._apply_data_validation(workbook)
                    
                    # Final atomic swap
                    os.replace(temp_path, self.excel_path)
                    self.logger.debug(f"Atomically saved tracker to {self.excel_path}")
                except Exception as e_save:
                    self.logger.error(f"Failed to write temporary tracker file: {e_save}")
                    if temp_path.exists(): os.remove(temp_path)
                    raise
            else:
                self.logger.error(f"Could not acquire lock for saving tracker: {self.excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to save tracker: {e}")
            raise
        finally:
            self._release_lock(lock_path)
            if temp_path.exists():
                try: os.remove(temp_path)
                except: pass

    def _ensure_row(self, sheet_name: str, subject_id: str, session: str, study: Optional[str] = None, extra_keys: Optional[Dict[str, Any]] = None) -> int:
        """
        Ensure a row exists for the subject/session/study and optional extra keys.
        Returns the index of the row.
        """
        if sheet_name not in self._data:
            # Create empty df with core columns
            cols = ['Subject_ID', 'Session']
            if study: cols.append('Study')
            if extra_keys:
                for k in extra_keys.keys():
                    if k not in cols: cols.append(k)
            self._data[sheet_name] = pd.DataFrame(columns=cols)
            
        df = self._data[sheet_name]
        
        # Normalize session for consistent matching
        def _normalize_session(s):
            if pd.isna(s) or s is None: return "N/A"
            s_str = str(s).strip()
            if s_str == "" or s_str.lower() == "none" or s_str == "nan": return "N/A"
            # Strip 'ses-' prefix for consistent storage
            if s_str.startswith("ses-"): s_str = s_str[4:]
            # Strip leading zeros for numeric comparison
            if s_str.isdigit(): s_str = str(int(s_str))
            return s_str if s_str else "N/A"
        
        session_norm = _normalize_session(session)
        
        # Build mask for matching (handling NaN safely)
        subj_mask = (df['Subject_ID'].astype(str) == str(subject_id))
        
        if session_norm == "N/A":
            ses_mask = df['Session'].isna() | (df['Session'].astype(str).apply(_normalize_session) == "N/A")
        else:
            ses_mask = df['Session'].astype(str).apply(_normalize_session) == session_norm
        
        mask = subj_mask & ses_mask
        
        if study and 'Study' in df.columns:
            mask = mask & (df['Study'] == study)
            
        if extra_keys:
            for k, v in extra_keys.items():
                if k in df.columns:
                    mask = mask & (df[k] == v)
                else:
                    # Column doesn't exist, so row definitely doesn't exist
                    mask = pd.Series([False] * len(df))
                    break
            
        matches = df.index[mask]
        
        if len(matches) > 0:
            return matches[0]
        else:
            # Create new row - use normalized session value
            new_row = {'Subject_ID': str(subject_id), 'Session': session_norm}
            if study: new_row['Study'] = study
            if extra_keys:
                new_row.update(extra_keys)
                
            # Ensure any new columns from extra_keys are added to the DF
            new_cols = [c for c in new_row.keys() if c not in df.columns]
            if new_cols:
                df = df.reindex(columns=list(df.columns) + new_cols)
                for c in new_cols: df[c] = df[c].astype(object)
            
            self._data[sheet_name] = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            return len(self._data[sheet_name]) - 1

    def update_status(self, subject_id: str, session: str, module: str, status: str, study: Optional[str] = None, modality: Optional[str] = None):
        """
        Update the status of a specific processing module.
        If modality is provided (e.g. 'Anatomical'), updates the modality-specific status sheet.
        """
        sheet_name = 'Processing_Status'
        if modality:
            sheet_name = f"{modality.capitalize()}_Status"
            if sheet_name not in self.CORE_SHEETS:
                self.logger.warning(f"Modality status sheet {sheet_name} not recognized. Using default.")
                sheet_name = 'Processing_Status'

        idx = self._ensure_row(sheet_name, subject_id, session, study)
        df = self._data[sheet_name]
        
        # Determine column name
        if sheet_name == 'Processing_Status':
             col = f"{module}_Status" if module else "Overall_Pipeline_Status"
        else:
             # For modality sheets, module might map to Preprocessing or Analysis
             if module.lower() in ['preprocessing', 'preproc']:
                 col = "Preprocessing"
             elif module.lower() in ['analysis', 'modeling']:
                 col = "Analysis"
             elif module.lower() in ['overall', 'total']:
                 col = "Overall_Status"
             else:
                 col = module # Fallback

        if col not in df.columns:
            # Ensure it's object dtype from the start to avoid float64 FutureWarning
            df = df.reindex(columns=list(df.columns) + [col])
            df[col] = df[col].astype(object)
        elif pd.api.types.is_numeric_dtype(df[col]):
            # If it exists but is numeric (common on reload from Excel), cast to object
            # to allow string statuses like 'Complete' or 'running'
            df[col] = df[col].astype(object)
            
        df.at[idx, col] = status
        
        # Update date
        date_col = 'Last_Processing_Date' if sheet_name == 'Processing_Status' else 'Last_Update'
        if date_col in df.columns:
            df.at[idx, date_col] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
        self._data[sheet_name] = df.copy()

    def add_metrics(self, subject_id: str, session: str, metrics: Dict[str, Any], study: Optional[str] = None):
        """Add QC or biological metrics to the tracker."""
        idx = self._ensure_row('Quality_Metrics', subject_id, session, study)
        df = self._data['Quality_Metrics']
        
        # Batch add new columns to avoid fragmentation warnings
        new_cols = [c for c in metrics.keys() if c not in df.columns]
        if new_cols:
            df = df.reindex(columns=list(df.columns) + new_cols)
            # Ensure new columns can handle various dtypes
            for c in new_cols: df[c] = df[c].astype(object)
            
        for key, val in metrics.items():
            if key in df.columns and not isinstance(val, (int, float, complex, np.number)) and val is not None:
                if pd.api.types.is_numeric_dtype(df[key]):
                    df[key] = df[key].astype(object)
            df.at[idx, key] = val
        
        # De-fragment
        self._data['Quality_Metrics'] = df.copy()

    def add_roi_stats(self, subject_id: str, session: str, tsv_path: Path, atlas_name: str, study: Optional[str] = None):
        """
        Parse an ROI stats TSV and update the tracker in a 'Wide-ROI' format.
        Each Atlas gets its own sheet (e.g., HarvardOxford_Metrics).
        ROIs become columns. Model, Metric, and Statistic are rows.
        """
        if not tsv_path.exists():
            return
            
        new_stats = pd.read_csv(tsv_path, sep='\t')
        
        # Ensure we have a 'model' column.
        if 'model' not in new_stats.columns:
            if 'metric' in new_stats.columns:
                def _infer_model(m):
                    if '_' in str(m): return str(m).rsplit('_', 1)[0]
                    return 'Other'
                new_stats['model'] = new_stats['metric'].apply(_infer_model)
                new_stats['metric'] = new_stats['metric'].apply(lambda x: str(x).rsplit('_', 1)[-1] if '_' in str(x) else x)
            else:
                new_stats['model'] = 'ROI_Stats'
        
        # Melt stats so that each Statistic (Mean, Median, Std) becomes a row
        stat_cols = [c for c in ['mean', 'median', 'std'] if c in new_stats.columns]
        if not stat_cols:
             return

        # Prepare for PIVOTING: we want ROI as columns
        # First melt stats columns
        melted = new_stats.melt(
            id_vars=['roi_name', 'model', 'metric'],
            value_vars=stat_cols,
            var_name='Statistic',
            value_name='Value'
        )
        melted['Statistic'] = melted['Statistic'].str.capitalize()
        
        # Normalize session
        session_norm = self._norm_ses(session)
        
        # Sheet name is now Atlas-specific to avoid row blowout
        # Sanitize atlas name for excel sheet
        safe_atlas = str(atlas_name).replace(' ', '_').replace('/', '_')[:20]
        sheet_name = f"{safe_atlas}_Metrics"
        
        # Pivot to WIDE format: ROI_Name becomes columns
        wide_df = melted.pivot_table(
            index=['model', 'metric', 'Statistic'],
            columns='roi_name',
            values='Value'
        ).reset_index()
        
        # Add metadata columns
        wide_df.insert(0, 'Subject_ID', str(subject_id))
        wide_df.insert(1, 'Session', session_norm)
        wide_df.insert(2, 'Study', study or '')
        
        # Rename 'model' to 'Model', 'metric' to 'Metric' for consistency
        wide_df = wide_df.rename(columns={'model': 'Model', 'metric': 'Metric'})
        
        # Merge with existing sheet data
        if sheet_name not in self._data:
            self._data[sheet_name] = wide_df
        else:
            df = self._data[sheet_name]
            # Concat and deduplicate by Subject/Session/Model/Metric/Statistic
            merged = pd.concat([df, wide_df], ignore_index=True)
            subset = ['Subject_ID', 'Session', 'Study', 'Model', 'Metric', 'Statistic']
            # Re-normalize just in case
            for col in subset:
                 if col in merged.columns: merged[col] = merged[col].astype(str).str.strip()
            
            self._data[sheet_name] = merged.drop_duplicates(subset=subset, keep='last')

    def update_metadata(self, subject_id: str, session: str, metadata: Dict[str, Any], study: Optional[str] = None):
        """Update subject demographic or scan metadata."""
        idx = self._ensure_row('Subject_Metadata', subject_id, session, study)
        df = self._data['Subject_Metadata']
        
        new_cols = [c for c in metadata.keys() if c not in df.columns]
        if new_cols:
            df = df.reindex(columns=list(df.columns) + new_cols)
            for c in new_cols: df[c] = df[c].astype(object)
            
        for key, val in metadata.items():
            if key in df.columns and not isinstance(val, (int, float, complex, np.number)) and val is not None:
                if pd.api.types.is_numeric_dtype(df[key]):
                    df[key] = df[key].astype(object)
            df.at[idx, key] = val
            
        self._data['Subject_Metadata'] = df.copy()
            
    def log_error(self, subject_id: str, session: str, module: str, error_msg: str, study: Optional[str] = None):
        """Log a processing error."""
        idx = self._ensure_row('Errors_Notes', subject_id, session, study)
        df = self._data['Errors_Notes']
        for col in ['Has_Processing_Errors', 'Error_Module', 'Error_Message', 'Error_Timestamp']:
             if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                  df[col] = df[col].astype(object)

        df.at[idx, 'Has_Processing_Errors'] = True
        df.at[idx, 'Error_Module'] = module
        df.at[idx, 'Error_Message'] = error_msg
        df.at[idx, 'Error_Timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def log_time(self, subject_id: str, session: str, module: str, time_min: float, study: Optional[str] = None):
        """Log processing time for a module."""
        idx = self._ensure_row('Processing_Times', subject_id, session, study)
        col_name = f"{module}_Time_Min"
        self._data['Processing_Times'].at[idx, col_name] = time_min

    def log_processing_step(self, subject_id: str, session: str, modality: str, step_name: str, 
                            details: Dict[str, Any], study: Optional[str] = None):
        """
        Log fine-grained processing step details (mirrors PDF report data).
        Each parameter in 'details' becomes a separate row for flexible querying.
        """
        from datetime import datetime
        
        sheet_name = 'Processing_Details'
        if sheet_name not in self._data:
            self._data[sheet_name] = pd.DataFrame(columns=[
                'Subject_ID', 'Session', 'Study', 'Modality', 'Step_Name', 'Parameter', 'Value', 'Timestamp'
            ])
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Normalize session
        def _norm_ses(s):
            if pd.isna(s) or s is None: return "N/A"
            s_str = str(s).strip().replace('ses-', '')
            if s_str.isdigit(): s_str = str(int(s_str))
            return s_str if s_str else "N/A"
        
        session_norm = _norm_ses(session)
        
        new_rows = []
        for param, value in details.items():
            # Skip some internal keys
            if param in ['figures', 'tables', 'Status'] and param != 'Status':
                continue
            new_rows.append({
                'Subject_ID': str(subject_id),
                'Session': session_norm,
                'Study': study or '',
                'Modality': modality,
                'Step_Name': step_name,
                'Parameter': param,
                'Value': str(value) if value is not None else '',
                'Timestamp': timestamp
            })
        
        if new_rows:
            # Check for existing entries and update (upsert by Subject_ID, Session, Study, Modality, Step_Name, Parameter)
            df = self._data[sheet_name]
            for row in new_rows:
                mask = (
                    (df['Subject_ID'].astype(str) == row['Subject_ID']) &
                    (df['Session'].astype(str) == row['Session']) &
                    (df['Study'] == row['Study']) &
                    (df['Modality'] == row['Modality']) &
                    (df['Step_Name'] == row['Step_Name']) &
                    (df['Parameter'] == row['Parameter'])
                )
                if mask.any():
                    idx = df.index[mask][0]
                    # Ensure columns can handle string values
                    for col in ['Value', 'Timestamp']:
                         if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                              df[col] = df[col].astype(object)
                    df.at[idx, 'Value'] = row['Value']
                    df.at[idx, 'Timestamp'] = row['Timestamp']
                else:
                    self._data[sheet_name] = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
                    df = self._data[sheet_name]

    def log_volume_statistics(
        self, 
        subject_id: str, 
        session: str, 
        volumes: Dict[str, float],
        method: str = 'freesurfer',
        icv: Optional[float] = None,
        study: Optional[str] = None
    ):
        """
        Log anatomical volume statistics to the tracker.
        
        Args:
            subject_id: Subject ID.
            session: Session ID.
            volumes: Dictionary of structure_name -> volume in mm³.
            method: Segmentation method ('freesurfer', 'fast', 'first').
            icv: Intracranial volume for normalization (optional).
            study: Study name.
        """
        sheet_name = 'Volume_Statistics'
        if sheet_name not in self._data:
            self._data[sheet_name] = pd.DataFrame(columns=[
                'Subject_ID', 'Session', 'Study', 'Method', 'Structure', 'Volume_mm3', 'ICV_Normalized'
            ])
        
        # Normalize session
        def _norm_ses(s):
            if pd.isna(s) or s is None: return "N/A"
            s_str = str(s).strip().replace('ses-', '')
            if s_str.isdigit(): s_str = str(int(s_str))
            return s_str if s_str else "N/A"
        
        session_norm = _norm_ses(session)
        
        new_rows = []
        for struct_name, vol_mm3 in volumes.items():
            # Compute ICV-normalized value if ICV provided
            icv_norm = vol_mm3 / icv if icv and icv > 0 else None
            
            new_rows.append({
                'Subject_ID': str(subject_id),
                'Session': session_norm,
                'Study': study or '',
                'Method': method,
                'Structure': struct_name.replace('_Volume_mm3', '').replace('_mm3', ''),
                'Volume_mm3': vol_mm3,
                'ICV_Normalized': icv_norm,
            })
        
        if new_rows:
            # Upsert logic: update existing or append
            df = self._data[sheet_name]
            for row in new_rows:
                mask = (
                    (df['Subject_ID'].astype(str) == row['Subject_ID']) &
                    (df['Session'].astype(str) == row['Session']) &
                    (df['Study'] == row['Study']) &
                    (df['Method'] == row['Method']) &
                    (df['Structure'] == row['Structure'])
                )
                if mask.any():
                    idx = df.index[mask][0]
                    df.at[idx, 'Volume_mm3'] = row['Volume_mm3']
                    df.at[idx, 'ICV_Normalized'] = row['ICV_Normalized']
                else:
                    self._data[sheet_name] = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
                    df = self._data[sheet_name]

    def log_roi_metrics(
        self,
        subject_id: str,
        session: str,
        roi_metrics: pd.DataFrame,
        study: Optional[str] = None
    ):
        """
        Log ROI metrics (cross-modal extraction) to the tracker in Wide-ROI format.
        Grouping by ROI_Source (Atlas) into separate sheets.
        """
        if roi_metrics.empty:
            return
            
        session_norm = self._norm_ses(session)
        
        # Group by Atlas
        for atlas, atlas_df in roi_metrics.groupby('ROI_Source'):
             safe_atlas = str(atlas).replace(' ', '_').replace('/', '_')[:20]
             sheet_name = f"{safe_atlas}_Metrics"
             
             # Pivot to WIDE format
             # If Modality is present, include it in index
             index_cols = ['Modality', 'Metric', 'Statistic']
             actual_index = [c for c in index_cols if c in atlas_df.columns]
             
             wide_df = atlas_df.pivot_table(
                 index=actual_index,
                 columns='ROI_Name',
                 values='Value'
             ).reset_index()
             
             # Add metadata
             wide_df.insert(0, 'Subject_ID', str(subject_id))
             wide_df.insert(1, 'Session', session_norm)
             wide_df.insert(2, 'Study', study or '')
             
             # Merge with existing
             if sheet_name not in self._data:
                 self._data[sheet_name] = wide_df
             else:
                 df = self._data[sheet_name]
                 merged = pd.concat([df, wide_df], ignore_index=True)
                 subset = ['Subject_ID', 'Session', 'Study'] + actual_index
                 self._data[sheet_name] = merged.drop_duplicates(subset=subset, keep='last')

    def _norm_ses(self, s):
        if pd.isna(s) or s is None: return "N/A"
        s_str = str(s).strip().replace('ses-', '')
        if s_str.isdigit(): s_str = str(int(s_str))
        return s_str if s_str else "N/A"

    def _recalculate_summary(self):
        """Recalculate high-level metrics for the Summary sheet."""
        rows = []
        try:
            if 'Processing_Status' in self._data:
                df = self._data['Processing_Status']
                if not df.empty and 'Subject_ID' in df.columns:
                    rows.append(['Total Subjects', int(df['Subject_ID'].nunique())])
                    rows.append(['Total Sessions', int(len(df))])
                    
                    # Completion rate
                    if 'Overall_Pipeline_Status' in df.columns:
                        # Handle cases where column might be missing or all NaN
                        complete = (df['Overall_Pipeline_Status'].astype(str).str.contains('Complete', case=False)).sum()
                        rate = (complete / len(df)) * 100 if len(df) > 0 else 0
                        rows.append(['Pipeline Completion Rate (%)', f"{rate:.1f}%"])
                
            # Add modality specific stats
            for mod in ['Anatomical', 'Diffusion', 'Relaxometry']:
                sheet = f"{mod}_Status"
                if sheet in self._data:
                    mdf = self._data[sheet]
                    if not mdf.empty and 'Overall_Status' in mdf.columns:
                        comp = (mdf['Overall_Status'].astype(str).str.contains('Complete', case=False)).sum()
                        rows.append([f'{mod} Completion', f"{comp}/{len(mdf)}"])
        except Exception as e:
            self.logger.warning(f"Error recalculating summary: {e}")

        if rows:
            self._data['Summary'] = pd.DataFrame(rows, columns=['Metric', 'Value'])

    def _apply_styles(self, workbook):
        """Apply conditional formatting and styling to status sheets using openpyxl."""
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Define styles
        styles = {
            'Complete': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'In Progress': PatternFill(start_color='BEE5EB', end_color='BEE5EB', fill_type='solid'),
            'Pending': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
            'Queued': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
            'Warning': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'Error': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'Failed': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        }
        
        # Text colors
        fonts = {
            'Complete': Font(color='006100', bold=True),
            'In Progress': Font(color='0C5460', bold=True),
            'Warning': Font(color='9C5700', bold=True),
            'Error': Font(color='9C0006', bold=True),
            'Failed': Font(color='9C0006', bold=True)
        }

        # Header style
        header_fill = PatternFill(start_color='4472CC', end_color='4472CC', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format Header
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

            # Apply Status Colors if it's a status sheet
            if 'Status' in sheet_name:
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if not cell.value: continue
                        val = str(cell.value)
                        for key, fill in styles.items():
                            if key.lower() in val.lower():
                                cell.fill = fill
                                if key in fonts:
                                    cell.font = fonts[key]
                                break
            
            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except: pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap at 50

    def _apply_data_validation(self, workbook):
        """Add data validation (dropdowns) to status columns using openpyxl."""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Valid statuses
        status_list = '"Complete,In Progress,Pending,Queued,Warning,Error,Failed,N/A,Manual Pass,Manual Fail"'
        dv = DataValidation(type="list", formula1=status_list, allow_blank=True)
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'Status Selection'

        for sheet_name in workbook.sheetnames:
            if 'Status' not in sheet_name:
                continue
                
            ws = workbook[sheet_name]
            # Identify status columns (usually everything except ID, Session, Study, Date)
            exclude_cols = ['Subject_ID', 'Session', 'Study', 'Last_Update', 'Last_Processing_Date', 'Segmentation_Method', 'B1_Mapping_Method', 'Atlases', 'Model_Fits']
            
            # Find columns to apply DV to
            dv_added = False
            if ws.max_row >= 1:
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and cell.value not in exclude_cols:
                        col_letter = cell.column_letter
                        # Apply to the rest of the column (up to row 1000 for convenience)
                        dv.add(f"{col_letter}2:{col_letter}1000")
                        dv_added = True
            
            if dv_added:
                ws.add_data_validation(dv)
